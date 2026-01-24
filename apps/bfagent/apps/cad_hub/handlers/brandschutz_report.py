"""
Brandschutz Prüfbericht Export.

Erstellt Prüfberichte in verschiedenen Formaten:
- PDF: Vollständiger Prüfbericht mit Grafiken
- Excel: Mängelliste und Checkliste
- JSON: Maschinenlesbare Daten
- HTML: Web-Ansicht

Inhalte:
- Zusammenfassung
- Mängelliste mit Prioritäten
- Symbol-Übersicht
- Fluchtweg-Analyse
- Regelwerk-Referenzen
- Empfehlungen
"""
import json
import logging
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional

from .base import (
    BaseCADHandler,
    CADHandlerResult,
    HandlerStatus,
)

logger = logging.getLogger(__name__)


@dataclass
class BerichtKonfiguration:
    """Konfiguration für Berichtserstellung."""
    titel: str = "Brandschutz-Prüfbericht"
    projekt_name: str = ""
    etage: str = ""
    pruefer: str = ""
    datum: str = ""
    
    # Inhalte
    mit_zusammenfassung: bool = True
    mit_maengelliste: bool = True
    mit_symboluebersicht: bool = True
    mit_fluchtweganalyse: bool = True
    mit_regelwerkreferenzen: bool = True
    mit_empfehlungen: bool = True
    mit_grafiken: bool = True
    
    # Ausgabeformat
    format: str = "pdf"  # pdf, excel, json, html


class BrandschutzReportHandler(BaseCADHandler):
    """
    Handler für Brandschutz-Prüfbericht-Erstellung.
    
    Erstellt professionelle Prüfberichte aus Analyseergebnissen.
    
    Input:
        analyse_ergebnis: dict - Ergebnis von BrandschutzHandler
        symbol_ergebnis: dict - Ergebnis von BrandschutzSymbolHandler (optional)
        konfiguration: dict - BerichtKonfiguration
        format: str - "pdf", "excel", "json", "html"
    
    Output:
        bericht: bytes - Generierter Bericht
        bericht_pfad: str - Pfad zur gespeicherten Datei (wenn save=True)
    """
    
    name = "BrandschutzReportHandler"
    description = "Erstellt Brandschutz-Prüfberichte"
    required_inputs = ["analyse_ergebnis"]
    optional_inputs = ["symbol_ergebnis", "konfiguration", "format", "save", "output_path"]
    
    def execute(self, input_data: dict) -> CADHandlerResult:
        """Erstellt Prüfbericht."""
        result = CADHandlerResult(
            success=True,
            handler_name=self.name,
            status=HandlerStatus.RUNNING,
        )
        
        analyse = input_data.get("analyse_ergebnis", {})
        symbole = input_data.get("symbol_ergebnis", {})
        config_dict = input_data.get("konfiguration", {})
        output_format = input_data.get("format", "html")
        save = input_data.get("save", False)
        output_path = input_data.get("output_path")
        
        # Konfiguration erstellen
        config = BerichtKonfiguration(
            projekt_name=config_dict.get("projekt_name", "Unbekannt"),
            etage=config_dict.get("etage", ""),
            pruefer=config_dict.get("pruefer", ""),
            datum=config_dict.get("datum", datetime.now().strftime("%d.%m.%Y")),
            format=output_format,
        )
        
        try:
            if output_format == "html":
                bericht_bytes, mime_type = self._generate_html(analyse, symbole, config)
            elif output_format == "json":
                bericht_bytes, mime_type = self._generate_json(analyse, symbole, config)
            elif output_format == "excel":
                bericht_bytes, mime_type = self._generate_excel(analyse, symbole, config)
            elif output_format == "pdf":
                bericht_bytes, mime_type = self._generate_pdf(analyse, symbole, config)
            else:
                result.add_error(f"Unbekanntes Format: {output_format}")
                return result
            
            # Speichern wenn gewünscht
            if save and output_path:
                Path(output_path).write_bytes(bericht_bytes)
                result.data["bericht_pfad"] = output_path
            
            result.data["bericht"] = bericht_bytes
            result.data["mime_type"] = mime_type
            result.data["format"] = output_format
            result.data["groesse_bytes"] = len(bericht_bytes)
            
        except Exception as e:
            result.add_error(f"Berichtserstellung fehlgeschlagen: {e}")
            logger.exception(f"[{self.name}] Fehler bei Berichtserstellung")
            return result
        
        result.status = HandlerStatus.SUCCESS
        logger.info(f"[{self.name}] Bericht erstellt: {output_format}, {len(bericht_bytes)} bytes")
        
        return result
    
    def _generate_html(self, analyse: dict, symbole: dict, config: BerichtKonfiguration) -> tuple[bytes, str]:
        """Generiert HTML-Bericht."""
        
        # Daten extrahieren
        brandschutz = analyse.get("brandschutz", {})
        zusammenfassung = brandschutz.get("zusammenfassung", {})
        maengel = brandschutz.get("maengel", [])
        warnungen = brandschutz.get("warnungen", [])
        fluchtwege = brandschutz.get("fluchtwege", [])
        einrichtungen = brandschutz.get("einrichtungen", [])
        
        symbol_stats = symbole.get("symbole", {}).get("statistik", {})
        vorgeschlagen = symbole.get("symbole", {}).get("vorgeschlagene_symbole", [])
        
        # HTML generieren
        html = f"""<!DOCTYPE html>
<html lang="de">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{config.titel}</title>
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: 'Segoe UI', Arial, sans-serif; line-height: 1.6; color: #333; padding: 40px; max-width: 1000px; margin: 0 auto; }}
        h1 {{ color: #c0392b; border-bottom: 3px solid #c0392b; padding-bottom: 10px; margin-bottom: 20px; }}
        h2 {{ color: #2c3e50; margin-top: 30px; margin-bottom: 15px; border-left: 4px solid #c0392b; padding-left: 10px; }}
        h3 {{ color: #34495e; margin-top: 20px; }}
        .header {{ background: linear-gradient(135deg, #c0392b, #e74c3c); color: white; padding: 30px; border-radius: 10px; margin-bottom: 30px; }}
        .header h1 {{ color: white; border: none; }}
        .meta {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 10px; margin-top: 15px; }}
        .meta-item {{ background: rgba(255,255,255,0.2); padding: 8px 15px; border-radius: 5px; }}
        .section {{ background: #f8f9fa; padding: 20px; border-radius: 8px; margin-bottom: 20px; }}
        .stats {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin: 20px 0; }}
        .stat-card {{ background: white; padding: 20px; border-radius: 8px; text-align: center; box-shadow: 0 2px 5px rgba(0,0,0,0.1); }}
        .stat-value {{ font-size: 2em; font-weight: bold; color: #c0392b; }}
        .stat-label {{ color: #666; font-size: 0.9em; }}
        table {{ width: 100%; border-collapse: collapse; margin: 15px 0; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background: #2c3e50; color: white; }}
        tr:hover {{ background: #f5f5f5; }}
        .badge {{ display: inline-block; padding: 4px 10px; border-radius: 15px; font-size: 0.85em; font-weight: bold; }}
        .badge-kritisch {{ background: #c0392b; color: white; }}
        .badge-hoch {{ background: #e67e22; color: white; }}
        .badge-mittel {{ background: #f1c40f; color: #333; }}
        .badge-gering {{ background: #27ae60; color: white; }}
        .badge-ok {{ background: #27ae60; color: white; }}
        .badge-warnung {{ background: #f39c12; color: white; }}
        .warning-box {{ background: #ffeaa7; border-left: 4px solid #f39c12; padding: 15px; margin: 15px 0; }}
        .error-box {{ background: #fab1a0; border-left: 4px solid #e74c3c; padding: 15px; margin: 15px 0; }}
        .success-box {{ background: #d4edda; border-left: 4px solid #27ae60; padding: 15px; margin: 15px 0; }}
        .footer {{ margin-top: 40px; padding-top: 20px; border-top: 1px solid #ddd; color: #666; font-size: 0.9em; }}
        @media print {{ body {{ padding: 20px; }} .header {{ -webkit-print-color-adjust: exact; }} }}
    </style>
</head>
<body>
    <div class="header">
        <h1>🔥 {config.titel}</h1>
        <div class="meta">
            <div class="meta-item"><strong>Projekt:</strong> {config.projekt_name}</div>
            <div class="meta-item"><strong>Etage:</strong> {config.etage or 'Alle'}</div>
            <div class="meta-item"><strong>Prüfer:</strong> {config.pruefer or 'N/A'}</div>
            <div class="meta-item"><strong>Datum:</strong> {config.datum}</div>
        </div>
    </div>
    
    <h2>📊 Zusammenfassung</h2>
    <div class="stats">
        <div class="stat-card">
            <div class="stat-value">{zusammenfassung.get('notausgaenge', 0)}</div>
            <div class="stat-label">Notausgänge</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">{zusammenfassung.get('feuerloescher', 0)}</div>
            <div class="stat-label">Feuerlöscher</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">{zusammenfassung.get('rauchmelder', 0)}</div>
            <div class="stat-label">Rauchmelder</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">{len(fluchtwege)}</div>
            <div class="stat-label">Fluchtwege</div>
        </div>
    </div>
"""
        
        # Mängel
        if maengel:
            html += f"""
    <h2>⚠️ Mängel ({len(maengel)})</h2>
    <div class="error-box">
        <strong>Folgende Mängel wurden festgestellt:</strong>
    </div>
    <table>
        <tr><th>#</th><th>Beschreibung</th><th>Schweregrad</th></tr>
"""
            for i, mangel in enumerate(maengel, 1):
                html += f"        <tr><td>{i}</td><td>{mangel}</td><td><span class='badge badge-kritisch'>Kritisch</span></td></tr>\n"
            html += "    </table>\n"
        else:
            html += """
    <h2>✅ Mängel</h2>
    <div class="success-box">
        <strong>Keine Mängel festgestellt.</strong>
    </div>
"""
        
        # Warnungen
        if warnungen:
            html += f"""
    <h2>⚡ Warnungen ({len(warnungen)})</h2>
    <div class="warning-box">
        <strong>Hinweise zur Beachtung:</strong>
    </div>
    <ul style="margin: 15px 0; padding-left: 20px;">
"""
            for warnung in warnungen:
                html += f"        <li>{warnung}</li>\n"
            html += "    </ul>\n"
        
        # Vorgeschlagene Symbole
        if vorgeschlagen:
            html += f"""
    <h2>➕ Empfohlene Ergänzungen ({len(vorgeschlagen)})</h2>
    <div class="section">
        <table>
            <tr><th>Symbol</th><th>Begründung</th><th>Priorität</th></tr>
"""
            for sym in vorgeschlagen[:10]:  # Max 10
                prio_class = "kritisch" if sym.get("prioritaet") == 1 else "mittel"
                html += f"""            <tr>
                <td>{sym.get('symbol_typ', '')}</td>
                <td>{sym.get('begruendung', '')}</td>
                <td><span class="badge badge-{prio_class}">P{sym.get('prioritaet', 2)}</span></td>
            </tr>
"""
            html += "        </table>\n    </div>\n"
            
            # Statistik der fehlenden Symbole
            if symbol_stats:
                html += """
    <div class="stats">
"""
                if symbol_stats.get("feuerloescher_fehlen", 0) > 0:
                    html += f"""        <div class="stat-card">
            <div class="stat-value" style="color: #e74c3c;">{symbol_stats['feuerloescher_fehlen']}</div>
            <div class="stat-label">Feuerlöscher fehlen</div>
        </div>
"""
                if symbol_stats.get("rauchmelder_fehlen", 0) > 0:
                    html += f"""        <div class="stat-card">
            <div class="stat-value" style="color: #e74c3c;">{symbol_stats['rauchmelder_fehlen']}</div>
            <div class="stat-label">Rauchmelder fehlen</div>
        </div>
"""
                html += "    </div>\n"
        
        # Footer
        html += f"""
    <div class="footer">
        <p>Erstellt am {datetime.now().strftime('%d.%m.%Y %H:%M')} | Brandschutz-Prüfsystem</p>
        <p><em>Dieser Bericht ersetzt keine behördliche Prüfung. Alle Angaben ohne Gewähr.</em></p>
    </div>
</body>
</html>"""
        
        return html.encode("utf-8"), "text/html"
    
    def _generate_json(self, analyse: dict, symbole: dict, config: BerichtKonfiguration) -> tuple[bytes, str]:
        """Generiert JSON-Bericht."""
        bericht = {
            "meta": {
                "titel": config.titel,
                "projekt": config.projekt_name,
                "etage": config.etage,
                "pruefer": config.pruefer,
                "datum": config.datum,
                "erstellt_am": datetime.now().isoformat(),
            },
            "analyse": analyse,
            "symbole": symbole,
        }
        
        return json.dumps(bericht, indent=2, ensure_ascii=False).encode("utf-8"), "application/json"
    
    def _generate_excel(self, analyse: dict, symbole: dict, config: BerichtKonfiguration) -> tuple[bytes, str]:
        """Generiert Excel-Bericht."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.warning(f"[{self.name}] openpyxl nicht installiert")
            # Fallback zu CSV
            return self._generate_csv_fallback(analyse, symbole, config)
        
        wb = openpyxl.Workbook()
        
        # Sheet 1: Zusammenfassung
        ws = wb.active
        ws.title = "Zusammenfassung"
        
        # Header-Style
        header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Titel
        ws["A1"] = config.titel
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")
        
        # Meta
        ws["A3"] = "Projekt:"
        ws["B3"] = config.projekt_name
        ws["A4"] = "Etage:"
        ws["B4"] = config.etage
        ws["A5"] = "Prüfer:"
        ws["B5"] = config.pruefer
        ws["A6"] = "Datum:"
        ws["B6"] = config.datum
        
        # Statistik
        brandschutz = analyse.get("brandschutz", {})
        zusammenfassung = brandschutz.get("zusammenfassung", {})
        
        ws["A8"] = "Statistik"
        ws["A8"].font = Font(bold=True)
        ws["A9"] = "Notausgänge"
        ws["B9"] = zusammenfassung.get("notausgaenge", 0)
        ws["A10"] = "Feuerlöscher"
        ws["B10"] = zusammenfassung.get("feuerloescher", 0)
        ws["A11"] = "Rauchmelder"
        ws["B11"] = zusammenfassung.get("rauchmelder", 0)
        
        # Sheet 2: Mängel
        ws2 = wb.create_sheet("Mängel")
        ws2.append(["#", "Beschreibung", "Schweregrad", "Status"])
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for i, mangel in enumerate(brandschutz.get("maengel", []), 1):
            ws2.append([i, mangel, "Kritisch", "Offen"])
        
        # Sheet 3: Empfehlungen
        ws3 = wb.create_sheet("Empfehlungen")
        ws3.append(["Symbol", "Position X", "Position Y", "Begründung", "Priorität"])
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for sym in symbole.get("symbole", {}).get("vorgeschlagene_symbole", []):
            ws3.append([
                sym.get("symbol_typ", ""),
                sym.get("position_x", 0),
                sym.get("position_y", 0),
                sym.get("begruendung", ""),
                sym.get("prioritaet", 2),
            ])
        
        # Spaltenbreiten
        for ws in [wb.active, ws2, ws3]:
            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        # Als Bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    def _generate_csv_fallback(self, analyse: dict, symbole: dict, config: BerichtKonfiguration) -> tuple[bytes, str]:
        """Fallback zu CSV wenn openpyxl nicht verfügbar."""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output, delimiter=";")
        
        writer.writerow([config.titel])
        writer.writerow(["Projekt", config.projekt_name])
        writer.writerow(["Datum", config.datum])
        writer.writerow([])
        
        writer.writerow(["Mängel"])
        for mangel in analyse.get("brandschutz", {}).get("maengel", []):
            writer.writerow([mangel])
        
        return output.getvalue().encode("utf-8"), "text/csv"
    
    def _generate_pdf(self, analyse: dict, symbole: dict, config: BerichtKonfiguration) -> tuple[bytes, str]:
        """Generiert PDF-Bericht."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        except ImportError:
            logger.warning(f"[{self.name}] reportlab nicht installiert, verwende HTML")
            # Fallback zu HTML
            html_bytes, _ = self._generate_html(analyse, symbole, config)
            return html_bytes, "text/html"
        
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Title'], textColor=colors.darkred)
        heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], textColor=colors.darkblue)
        
        story = []
        
        # Titel
        story.append(Paragraph(f"🔥 {config.titel}", title_style))
        story.append(Spacer(1, 0.5*cm))
        
        # Meta
        meta_data = [
            ["Projekt:", config.projekt_name, "Etage:", config.etage or "Alle"],
            ["Prüfer:", config.pruefer or "N/A", "Datum:", config.datum],
        ]
        meta_table = Table(meta_data, colWidths=[3*cm, 5*cm, 3*cm, 5*cm])
        meta_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 1*cm))
        
        # Zusammenfassung
        story.append(Paragraph("Zusammenfassung", heading_style))
        brandschutz = analyse.get("brandschutz", {})
        zusammenfassung = brandschutz.get("zusammenfassung", {})
        
        stats_data = [
            ["Notausgänge", "Feuerlöscher", "Rauchmelder", "Fluchtwege"],
            [
                str(zusammenfassung.get("notausgaenge", 0)),
                str(zusammenfassung.get("feuerloescher", 0)),
                str(zusammenfassung.get("rauchmelder", 0)),
                str(len(brandschutz.get("fluchtwege", []))),
            ],
        ]
        stats_table = Table(stats_data, colWidths=[4*cm]*4)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkred),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 18),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 1*cm))
        
        # Mängel
        maengel = brandschutz.get("maengel", [])
        story.append(Paragraph(f"Mängel ({len(maengel)})", heading_style))
        if maengel:
            maengel_data = [["#", "Beschreibung"]]
            for i, m in enumerate(maengel, 1):
                maengel_data.append([str(i), m])
            maengel_table = Table(maengel_data, colWidths=[1*cm, 15*cm])
            maengel_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.orange),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(maengel_table)
        else:
            story.append(Paragraph("✅ Keine Mängel festgestellt.", styles['Normal']))
        
        doc.build(story)
        output.seek(0)
        
        return output.getvalue(), "application/pdf"


# Singleton
_report_handler: Optional[BrandschutzReportHandler] = None

def get_brandschutz_report_handler() -> BrandschutzReportHandler:
    """Gibt BrandschutzReportHandler-Instanz zurück."""
    global _report_handler
    if _report_handler is None:
        _report_handler = BrandschutzReportHandler()
    return _report_handler
