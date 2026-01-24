# apps/cad_hub/services/avb_service.py
"""
AVB Service - Ausschreibung, Vergabe, Bauausführung
===================================================

Service für kompletten Ausschreibungs- und Vergabeprozess.

Funktionen:
- LV-Erstellung aus IFC-Daten
- GAEB Export (X81, X83, X85)
- Angebotsvergleich (Preisspiegel)
- Vergabevorschlag
"""

import logging
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List, Optional

from django.db import transaction
from django.utils import timezone

logger = logging.getLogger(__name__)


@dataclass
class BidComparison:
    """Angebotsvergleich für eine Position"""
    position_oz: str
    position_text: str
    quantity: Decimal
    unit: str
    bids: List[Dict]  # [{"bidder": str, "unit_price": Decimal, "total": Decimal, "rank": int}]
    lowest_price: Decimal
    average_price: Decimal
    spread_percent: float  # Preisspreizung in %


@dataclass
class PriceRanking:
    """Preisranking eines Bieters"""
    bidder_id: str
    bidder_name: str
    total_price: Decimal
    final_price: Decimal  # Nach Rabatten
    rank: int
    price_score: Decimal  # 100 = günstigster
    positions_count: int
    lowest_positions: int  # Anzahl Positionen wo günstigster


class AVBService:
    """
    Service für Ausschreibungs-, Vergabe- und Bauausführungs-Prozesse.
    
    Usage:
        service = AVBService()
        tender = service.create_tender_from_ifc(ifc_model, "Rohbauarbeiten")
        comparison = service.compare_bids(tender)
        ranking = service.calculate_price_ranking(tender)
    """
    
    def create_tender_from_ifc(
        self,
        ifc_model,
        trade: str,
        cost_group: str = "",
        title: str = "",
        gewerke: Optional[List[str]] = None,
    ):
        """
        Erstellt Ausschreibung aus IFC-Modell.
        
        Args:
            ifc_model: IFCModel instance
            trade: Gewerk (z.B. "Rohbauarbeiten")
            cost_group: DIN 276 Kostengruppe
            title: Ausschreibungstitel
            gewerke: Liste der Gewerke für X83 Converter
            
        Returns:
            Tender instance
        """
        from ..models import ConstructionProject, Tender, TenderPosition
        from .ifc_x83_converter import get_ifc_x83_converter
        
        # Projekt ermitteln oder erstellen
        project, _ = ConstructionProject.objects.get_or_create(
            ifc_project=ifc_model.project,
            defaults={
                "client": "N/A",
                "project_number": f"P-{ifc_model.project.pk.hex[:8].upper()}"
            }
        )
        
        # Ausschreibungsnummer generieren
        tender_count = project.tenders.count() + 1
        tender_number = f"{project.project_number}-LV{tender_count:02d}"
        
        # IFC-Daten extrahieren
        converter = get_ifc_x83_converter()
        ifc_data = self._extract_ifc_data(ifc_model)
        
        # LV erstellen
        lv = converter._create_leistungsverzeichnis(
            ifc_data=ifc_data,
            projekt_name=ifc_model.project.name,
            projekt_nummer=tender_number,
            include_prices=False,
            selected_gewerke=gewerke,
        )
        
        with transaction.atomic():
            # Tender erstellen
            tender = Tender.objects.create(
                project=project,
                tender_number=tender_number,
                title=title or f"{trade} - {ifc_model.project.name}",
                trade=trade,
                cost_group=cost_group,
                status="draft",
                estimated_value=lv.netto_summe,
            )
            
            # Positionen aus LV übernehmen
            order = 0
            for los in lv.lose:
                for pos in los.positionen:
                    order += 1
                    TenderPosition.objects.create(
                        tender=tender,
                        oz=pos.oz,
                        short_text=pos.kurztext,
                        long_text=pos.langtext,
                        quantity=pos.menge,
                        unit=pos.einheit.value,
                        stlb_code=pos.stlb_code,
                        order=order,
                    )
            
            logger.info(f"Tender {tender_number} erstellt mit {order} Positionen")
        
        return tender
    
    def _extract_ifc_data(self, ifc_model) -> dict:
        """Extrahiert IFC-Daten aus der Datenbank."""
        from ..models import Door, Room, Slab, Wall, Window
        
        rooms = list(Room.objects.filter(ifc_model=ifc_model).values(
            "name", "number", "area", "perimeter", "height", "volume"
        ))
        
        walls = list(Wall.objects.filter(ifc_model=ifc_model).values(
            "name", "ifc_guid", "length", "height", "thickness"
        ))
        for wall in walls:
            wall["area"] = (wall.get("length", 0) or 0) * (wall.get("height", 0) or 0)
        
        doors = list(Door.objects.filter(ifc_model=ifc_model).values(
            "name", "ifc_guid", "width", "height"
        ))
        for door in doors:
            door["type"] = "Standard"
        
        windows = list(Window.objects.filter(ifc_model=ifc_model).values(
            "name", "ifc_guid", "width", "height"
        ))
        
        slabs = list(Slab.objects.filter(ifc_model=ifc_model).values(
            "name", "ifc_guid", "area", "thickness"
        ))
        
        return {
            "rooms": rooms,
            "walls": walls,
            "doors": doors,
            "windows": windows,
            "slabs": slabs,
        }
    
    def compare_bids(self, tender) -> List[BidComparison]:
        """
        Erstellt Preisspiegel / Angebotsvergleich.
        
        Args:
            tender: Tender instance
            
        Returns:
            Liste von BidComparison pro Position
        """
        from ..models import Bid, BidPosition, BidStatus
        
        # Eingegangene Angebote
        bids = tender.bids.filter(
            status__in=[BidStatus.RECEIVED, BidStatus.EVALUATED, BidStatus.NEGOTIATION]
        )
        
        if not bids.exists():
            return []
        
        comparisons = []
        
        for position in tender.positions.all():
            bid_data = []
            prices = []
            
            for bid in bids:
                try:
                    bp = BidPosition.objects.get(bid=bid, tender_position=position)
                    prices.append(bp.unit_price)
                    bid_data.append({
                        "bidder": bid.bidder.company_name,
                        "bidder_id": str(bid.bidder.pk),
                        "unit_price": bp.unit_price,
                        "total": bp.total_price,
                        "rank": 0,  # Wird später berechnet
                    })
                except BidPosition.DoesNotExist:
                    continue
            
            if not bid_data:
                continue
            
            # Ränge berechnen
            bid_data.sort(key=lambda x: x["unit_price"])
            for i, bd in enumerate(bid_data, 1):
                bd["rank"] = i
            
            lowest = min(prices)
            highest = max(prices)
            average = sum(prices) / len(prices)
            spread = ((highest - lowest) / lowest * 100) if lowest > 0 else 0
            
            comparisons.append(BidComparison(
                position_oz=position.oz,
                position_text=position.short_text,
                quantity=position.quantity,
                unit=position.unit,
                bids=bid_data,
                lowest_price=lowest,
                average_price=average,
                spread_percent=float(spread),
            ))
        
        return comparisons
    
    def calculate_price_ranking(self, tender) -> List[PriceRanking]:
        """
        Berechnet Preisranking aller Bieter.
        
        Args:
            tender: Tender instance
            
        Returns:
            Liste von PriceRanking, sortiert nach Endpreis
        """
        from ..models import Bid, BidStatus
        
        bids = tender.bids.filter(
            status__in=[BidStatus.RECEIVED, BidStatus.EVALUATED, BidStatus.NEGOTIATION]
        ).select_related('bidder')
        
        if not bids.exists():
            return []
        
        # Vergleich erstellen für "günstigste Positionen" Zählung
        comparisons = self.compare_bids(tender)
        
        rankings = []
        lowest_total = None
        
        for bid in bids.order_by('total_price'):
            # Zähle Positionen wo dieser Bieter günstigster ist
            lowest_count = 0
            for comp in comparisons:
                for bd in comp.bids:
                    if bd["bidder_id"] == str(bid.bidder.pk) and bd["rank"] == 1:
                        lowest_count += 1
                        break
            
            final = bid.final_price
            if lowest_total is None:
                lowest_total = final
            
            # Preisscore: 100 für günstigsten, linear abnehmend
            price_score = Decimal("100") * lowest_total / final if final > 0 else Decimal("0")
            
            rankings.append(PriceRanking(
                bidder_id=str(bid.bidder.pk),
                bidder_name=bid.bidder.company_name,
                total_price=bid.total_price,
                final_price=final,
                rank=len(rankings) + 1,
                price_score=price_score.quantize(Decimal("0.1")),
                positions_count=bid.positions.count(),
                lowest_positions=lowest_count,
            ))
        
        return rankings
    
    def suggest_award(self, tender) -> Optional[Dict]:
        """
        Erstellt Vergabevorschlag basierend auf Preis und Bewertung.
        
        Args:
            tender: Tender instance
            
        Returns:
            Dict mit Vergabevorschlag oder None
        """
        rankings = self.calculate_price_ranking(tender)
        
        if not rankings:
            return None
        
        # Einfacher Fall: Günstigster Bieter
        winner = rankings[0]
        
        return {
            "recommended_bidder": winner.bidder_name,
            "recommended_bidder_id": winner.bidder_id,
            "contract_value": winner.final_price,
            "savings_vs_estimate": tender.estimated_value - winner.final_price,
            "savings_percent": float(
                (tender.estimated_value - winner.final_price) / tender.estimated_value * 100
            ) if tender.estimated_value > 0 else 0,
            "ranking": rankings,
            "reason": "Günstigster Bieter nach Endpreis",
        }
    
    def export_tender_gaeb(
        self,
        tender,
        phase: str = "X81",
        include_prices: bool = False,
    ) -> BytesIO:
        """
        Exportiert Ausschreibung als GAEB.
        
        Args:
            tender: Tender instance
            phase: GAEB Phase (X81, X83, X84, X85)
            include_prices: Preise inkludieren (für X83+)
            
        Returns:
            BytesIO mit GAEB XML
        """
        from .gaeb_generator import (
            GAEBGenerator,
            GAEBPhase,
            Leistungsverzeichnis,
            LosGruppe,
            Position,
        )
        
        # LV erstellen
        positionen = []
        for pos in tender.positions.all():
            positionen.append(Position(
                oz=pos.oz,
                kurztext=pos.short_text,
                langtext=pos.long_text,
                menge=pos.quantity,
                einheit=pos.unit,
                stlb_code=pos.stlb_code,
            ))
        
        lv = Leistungsverzeichnis(
            projekt_name=tender.project.ifc_project.name,
            projekt_nummer=tender.tender_number,
            auftraggeber=tender.project.client,
            lose=[LosGruppe(oz="01", bezeichnung=tender.trade, positionen=positionen)],
            phase=GAEBPhase(phase[-2:]),  # "X81" -> "81"
        )
        
        generator = GAEBGenerator()
        return generator.generate_xml(lv)
    
    def export_price_comparison_excel(self, tender) -> BytesIO:
        """
        Exportiert Preisspiegel als Excel.
        
        Args:
            tender: Tender instance
            
        Returns:
            BytesIO mit Excel-Datei
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        
        comparisons = self.compare_bids(tender)
        rankings = self.calculate_price_ranking(tender)
        
        wb = Workbook()
        
        # Sheet 1: Preisspiegel
        ws1 = wb.active
        ws1.title = "Preisspiegel"
        
        # Header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="003366", fill_type="solid")
        
        headers = ["OZ", "Kurztext", "Menge", "Einheit"]
        bidder_names = [r.bidder_name for r in rankings]
        headers.extend(bidder_names)
        headers.extend(["Günstigster", "Durchschnitt", "Spreizung %"])
        
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        # Daten
        for row, comp in enumerate(comparisons, 2):
            ws1.cell(row=row, column=1, value=comp.position_oz)
            ws1.cell(row=row, column=2, value=comp.position_text)
            ws1.cell(row=row, column=3, value=float(comp.quantity))
            ws1.cell(row=row, column=4, value=comp.unit)
            
            # Bieterpreise
            col = 5
            for bidder_name in bidder_names:
                price = None
                for bd in comp.bids:
                    if bd["bidder"] == bidder_name:
                        price = float(bd["unit_price"])
                        break
                ws1.cell(row=row, column=col, value=price or "-")
                col += 1
            
            ws1.cell(row=row, column=col, value=float(comp.lowest_price))
            ws1.cell(row=row, column=col + 1, value=float(comp.average_price))
            ws1.cell(row=row, column=col + 2, value=f"{comp.spread_percent:.1f}%")
        
        # Sheet 2: Ranking
        ws2 = wb.create_sheet("Ranking")
        ranking_headers = ["Rang", "Bieter", "Netto", "Endpreis", "Score", "Günstigste Pos."]
        for col, h in enumerate(ranking_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for row, r in enumerate(rankings, 2):
            ws2.cell(row=row, column=1, value=r.rank)
            ws2.cell(row=row, column=2, value=r.bidder_name)
            ws2.cell(row=row, column=3, value=float(r.total_price))
            ws2.cell(row=row, column=4, value=float(r.final_price))
            ws2.cell(row=row, column=5, value=float(r.price_score))
            ws2.cell(row=row, column=6, value=r.lowest_positions)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# Singleton
_avb_service = None

def get_avb_service() -> AVBService:
    """Get singleton instance."""
    global _avb_service
    if _avb_service is None:
        _avb_service = AVBService()
    return _avb_service
