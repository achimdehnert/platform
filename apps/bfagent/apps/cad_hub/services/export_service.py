# apps/cad_hub/services/export_service.py
"""
Raumbuch Export Service

Basiert auf BauCAD Hub MCP generators/raumbuch.py und ifc_mcp Excel patterns.
Integriert DIN 277 Calculator für korrekte Flächenklassifizierung.
"""
import logging
from io import BytesIO
from pathlib import Path
from typing import Optional

from django.db.models import Sum

from .din277_calculator import DIN277Calculator, DIN277Result

logger = logging.getLogger(__name__)


class RaumbuchExportService:
    """
    Professioneller Raumbuch-Export nach Excel

    Features:
    - DIN 277 Klassifizierung
    - Geschoss-Gruppierung
    - Flächenübersicht
    - Professionelle Formatierung
    """

    def __init__(self):
        self.din277 = DIN277Calculator()
        self._setup_styles()

    def _setup_styles(self):
        """Initialisiert Excel-Styles"""
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        # Header
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.header_align = Alignment(horizontal="center", vertical="center")

        # Subheader (Geschosse)
        self.subheader_font = Font(bold=True, size=10)
        self.subheader_fill = PatternFill(
            start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
        )

        # Daten
        self.data_font = Font(size=10)
        self.number_align = Alignment(horizontal="right")

        # Border
        self.thin_border = Border(
            left=Side(style="thin", color="B4B4B4"),
            right=Side(style="thin", color="B4B4B4"),
            top=Side(style="thin", color="B4B4B4"),
            bottom=Side(style="thin", color="B4B4B4"),
        )

        # Summen
        self.sum_font = Font(bold=True, size=10)
        self.sum_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    def export_to_excel(self, ifc_model, include_din277: bool = True) -> BytesIO:
        """
        Exportiert komplettes Raumbuch nach Excel

        Sheets:
        1. Raumbuch - Alle Räume mit Details
        2. DIN 277 - Flächenübersicht (optional)
        3. Geschosse - Zusammenfassung pro Geschoss
        """
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        from ..models import Floor, Room

        wb = Workbook()

        # Sheet 1: Raumbuch
        ws_rooms = wb.active
        ws_rooms.title = "Raumbuch"
        self._create_raumbuch_sheet(ws_rooms, ifc_model)

        # Sheet 2: DIN 277 (optional)
        if include_din277:
            ws_din277 = wb.create_sheet("DIN 277")
            self._create_din277_sheet(ws_din277, ifc_model)

        # Sheet 3: Geschosse
        ws_floors = wb.create_sheet("Geschosse")
        self._create_floor_summary_sheet(ws_floors, ifc_model)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _create_raumbuch_sheet(self, ws, ifc_model):
        """Erstellt das Raumbuch-Sheet"""
        from openpyxl.utils import get_column_letter

        from ..models import Floor, Room

        # Titel
        ws["A1"] = f"Raumbuch - {ifc_model.project.name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:G1")

        ws["A2"] = f"Version {ifc_model.version} | {ifc_model.ifc_schema}"
        ws["A2"].font = Font(size=9, italic=True, color="666666")

        # Header (Zeile 4)
        headers = [
            ("Nr.", 10),
            ("Raumname", 25),
            ("Geschoss", 15),
            ("Fläche [m²]", 12),
            ("Höhe [m]", 10),
            ("Volumen [m³]", 12),
            ("DIN 277", 18),
        ]

        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_align
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col)].width = width

        # Räume nach Geschoss gruppiert
        floors = Floor.objects.filter(ifc_model=ifc_model).order_by("sort_order")
        row = 5

        for floor in floors:
            # Geschoss-Header
            ws.cell(row=row, column=1, value=floor.name).font = self.subheader_font
            ws.merge_cells(f"A{row}:G{row}")
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = self.subheader_fill
                ws.cell(row=row, column=col).border = self.thin_border
            row += 1

            # Räume des Geschosses
            rooms = Room.objects.filter(ifc_model=ifc_model, floor=floor).order_by("number")

            for room in rooms:
                din277 = self.din277.classify_room(room.name, room.usage_category)

                ws.cell(row=row, column=1, value=room.number).border = self.thin_border
                ws.cell(row=row, column=2, value=room.name).border = self.thin_border
                ws.cell(row=row, column=3, value=floor.name).border = self.thin_border

                area_cell = ws.cell(row=row, column=4, value=room.area)
                area_cell.number_format = "#,##0.00"
                area_cell.alignment = self.number_align
                area_cell.border = self.thin_border

                height_cell = ws.cell(row=row, column=5, value=room.height)
                height_cell.number_format = "0.00"
                height_cell.alignment = self.number_align
                height_cell.border = self.thin_border

                vol_cell = ws.cell(row=row, column=6, value=room.volume)
                vol_cell.number_format = "#,##0.00"
                vol_cell.alignment = self.number_align
                vol_cell.border = self.thin_border

                ws.cell(row=row, column=7, value=din277.value).border = self.thin_border
                row += 1

            # Geschoss-Summe
            floor_sum = rooms.aggregate(total=Sum("area"))["total"] or 0
            ws.cell(row=row, column=3, value="Summe").font = self.sum_font
            sum_cell = ws.cell(row=row, column=4, value=floor_sum)
            sum_cell.font = self.sum_font
            sum_cell.fill = self.sum_fill
            sum_cell.number_format = "#,##0.00"
            row += 1

        # Gesamtsumme
        row += 1
        total = Room.objects.filter(ifc_model=ifc_model).aggregate(total=Sum("area"))["total"] or 0

        ws.cell(row=row, column=3, value="GESAMT").font = Font(bold=True, size=11)
        total_cell = ws.cell(row=row, column=4, value=total)
        total_cell.font = Font(bold=True, size=11)
        total_cell.number_format = "#,##0.00"
        total_cell.fill = self.sum_fill

    def _create_din277_sheet(self, ws, ifc_model):
        """Erstellt DIN 277 Übersicht"""
        from ..models import Room

        # Berechnung
        rooms = Room.objects.filter(ifc_model=ifc_model)
        room_list = list(rooms.values("name", "area", "usage_category"))
        result = self.din277.calculate_from_rooms(room_list)

        # Titel
        ws["A1"] = "DIN 277:2021 Flächenberechnung"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = ifc_model.project.name
        ws["A2"].font = Font(size=10, italic=True)

        # Tabelle
        data = [
            ("BGF", "Brutto-Grundfläche", float(result.bgf)),
            ("KGF", "Konstruktions-Grundfläche", float(result.kgf)),
            ("NRF", "Netto-Raumfläche", float(result.nrf)),
            ("", "", ""),
            ("NF", "Nutzfläche", float(result.nf)),
            ("NF 1", "  Wohnen und Aufenthalt", float(result.nf1)),
            ("NF 2", "  Büroarbeit", float(result.nf2)),
            ("NF 3", "  Produktion/Küche", float(result.nf3)),
            ("NF 4", "  Lagern/Verkaufen", float(result.nf4)),
            ("NF 5", "  Bildung/Kultur", float(result.nf5)),
            ("NF 6", "  Heilen/Pflegen", float(result.nf6)),
            ("NF 7", "  Sonstige", float(result.nf7)),
            ("", "", ""),
            ("TF", "Technische Funktionsfläche", float(result.tf)),
            ("VF", "Verkehrsfläche", float(result.vf)),
            ("", "", ""),
            ("BRI", "Brutto-Rauminhalt", float(result.bri)),
        ]

        # Header
        ws["A4"] = "Kürzel"
        ws["B4"] = "Bezeichnung"
        ws["C4"] = "Fläche [m²]"
        for col in ["A", "B", "C"]:
            ws[f"{col}4"].font = self.header_font
            ws[f"{col}4"].fill = self.header_fill

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15

        # Daten
        for row, (code, name, value) in enumerate(data, 5):
            ws.cell(row=row, column=1, value=code)
            ws.cell(row=row, column=2, value=name)
            if value:
                val_cell = ws.cell(row=row, column=3, value=value)
                val_cell.number_format = "#,##0.00"
                val_cell.alignment = self.number_align

            # Hervorhebung Hauptkategorien
            if code in ["BGF", "NRF", "NF", "BRI"]:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).font = self.sum_font

        # Kennzahlen
        row = len(data) + 7
        ws.cell(row=row, column=1, value="Kennzahlen").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="NRF/BGF")
        ws.cell(row=row, column=2, value="Flächeneffizienz")
        ws.cell(row=row, column=3, value=result.nrf_ratio).number_format = "0.0%"
        row += 1
        ws.cell(row=row, column=1, value="VF/NRF")
        ws.cell(row=row, column=2, value="Verkehrsflächenanteil")
        ws.cell(row=row, column=3, value=result.vf_ratio).number_format = "0.0%"

    def _create_floor_summary_sheet(self, ws, ifc_model):
        """Erstellt Geschoss-Zusammenfassung"""
        from django.db.models import Count

        from ..models import Floor, Room

        ws["A1"] = "Geschossübersicht"
        ws["A1"].font = Font(bold=True, size=14)

        # Header
        headers = ["Geschoss", "Räume", "Fläche [m²]", "Anteil"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header).font = self.header_font
            ws.cell(row=3, column=col).fill = self.header_fill

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 10

        # Daten
        floors = (
            Floor.objects.filter(ifc_model=ifc_model)
            .annotate(room_count=Count("rooms"))
            .order_by("sort_order")
        )

        total_area = (
            Room.objects.filter(ifc_model=ifc_model).aggregate(total=Sum("area"))["total"] or 1
        )

        row = 4
        for floor in floors:
            floor_area = (
                Room.objects.filter(ifc_model=ifc_model, floor=floor).aggregate(total=Sum("area"))[
                    "total"
                ]
                or 0
            )

            ws.cell(row=row, column=1, value=floor.name)
            ws.cell(row=row, column=2, value=floor.room_count)
            ws.cell(row=row, column=3, value=floor_area).number_format = "#,##0.00"
            ws.cell(row=row, column=4, value=floor_area / total_area).number_format = "0.0%"
            row += 1

    def export_din277_summary(self, ifc_model) -> BytesIO:
        """Exportiert nur DIN 277 Übersicht"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "DIN 277"

        self._create_din277_sheet(ws, ifc_model)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
