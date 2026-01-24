# apps/cad_hub/services/gaeb_generator.py
"""
GAEB X84 Generator - Leistungsverzeichnis Export

Basiert auf BauCAD Hub MCP generators/gaeb.py
Für deutsche Bauausschreibungen nach GAEB Standard
"""
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal
from enum import Enum
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional


class GAEBPhase(str, Enum):
    """GAEB Datenaustausch-Phasen"""

    X81 = "81"  # Anfrage
    X83 = "83"  # Angebot (mit Preisen)
    X84 = "84"  # Nebenangebot
    X85 = "85"  # Auftragserteilung


class MengenEinheit(str, Enum):
    """Mengeneinheiten nach GAEB"""

    STK = "Stk"  # Stück
    M = "m"  # Meter
    M2 = "m2"  # Quadratmeter
    M3 = "m3"  # Kubikmeter
    KG = "kg"  # Kilogramm
    T = "t"  # Tonne
    L = "l"  # Liter
    H = "h"  # Stunde
    PAU = "psch"  # Pauschal
    TAG = "Tag"  # Tag


@dataclass
class Position:
    """LV-Position"""

    oz: str  # Ordnungszahl (z.B. "01.02.0010")
    kurztext: str
    langtext: str = ""
    menge: Decimal = Decimal("0")
    einheit: MengenEinheit = MengenEinheit.STK
    einheitspreis: Decimal = Decimal("0")
    gesamtpreis: Decimal = Decimal("0")
    stlb_code: str = ""  # STLB-Bau Referenz

    def __post_init__(self):
        if self.gesamtpreis == 0 and self.menge > 0 and self.einheitspreis > 0:
            self.gesamtpreis = self.menge * self.einheitspreis

    def to_dict(self) -> Dict:
        return {
            "oz": self.oz,
            "kurztext": self.kurztext,
            "menge": float(self.menge),
            "einheit": self.einheit.value,
            "einheitspreis": float(self.einheitspreis),
            "gesamtpreis": float(self.gesamtpreis),
        }


@dataclass
class LosGruppe:
    """Los/Titel/Gruppe im LV"""

    oz: str
    bezeichnung: str
    positionen: List[Position] = field(default_factory=list)
    untergruppen: List["LosGruppe"] = field(default_factory=list)

    @property
    def summe(self) -> Decimal:
        pos_summe = sum((p.gesamtpreis for p in self.positionen), Decimal("0"))
        ug_summe = sum((ug.summe for ug in self.untergruppen), Decimal("0"))
        return pos_summe + ug_summe

    @property
    def anzahl_positionen(self) -> int:
        return len(self.positionen) + sum(ug.anzahl_positionen for ug in self.untergruppen)


@dataclass
class Leistungsverzeichnis:
    """Komplettes Leistungsverzeichnis"""

    projekt_name: str
    projekt_nummer: str = ""
    lv_nummer: str = ""
    auftraggeber: str = ""
    auftragnehmer: str = ""
    lose: List[LosGruppe] = field(default_factory=list)
    waehrung: str = "EUR"
    datum: date = field(default_factory=date.today)
    phase: GAEBPhase = GAEBPhase.X83

    @property
    def netto_summe(self) -> Decimal:
        return sum((los.summe for los in self.lose), Decimal("0"))

    @property
    def mwst(self) -> Decimal:
        return self.netto_summe * Decimal("0.19")

    @property
    def brutto_summe(self) -> Decimal:
        return self.netto_summe + self.mwst

    @property
    def anzahl_positionen(self) -> int:
        return sum(los.anzahl_positionen for los in self.lose)


class GAEBGenerator:
    """
    Generiert GAEB X84 XML und Excel

    Basiert auf BauCAD Hub MCP GAEBX84Generator
    """

    GAEB_NAMESPACE = "http://www.gaeb.de/GAEB_DA_XML/200407"

    def generate_xml(self, lv: Leistungsverzeichnis) -> BytesIO:
        """Generiert GAEB X84 XML"""
        root = self._create_root(lv)
        self._indent(root)

        output = BytesIO()
        tree = ET.ElementTree(root)
        tree.write(output, encoding="utf-8", xml_declaration=True)
        output.seek(0)
        return output

    def generate_excel(self, lv: Leistungsverzeichnis) -> BytesIO:
        """Generiert Excel-Alternative zum GAEB"""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Leistungsverzeichnis"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        sum_font = Font(bold=True)
        sum_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # Projekt-Info
        ws["A1"] = f"Leistungsverzeichnis: {lv.projekt_name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Projekt-Nr.: {lv.projekt_nummer}" if lv.projekt_nummer else ""
        ws["A3"] = f"Datum: {lv.datum.strftime('%d.%m.%Y')}"

        # Header (Zeile 5)
        headers = ["OZ", "Kurztext", "Menge", "Einheit", "EP [€]", "GP [€]"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12

        # Daten
        row = 6
        for los in lv.lose:
            row = self._write_gruppe_excel(ws, los, row, sum_font, sum_fill)

        # Summen
        row += 1
        ws.cell(row=row, column=2, value="NETTO SUMME:").font = sum_font
        ws.cell(row=row, column=6, value=float(lv.netto_summe)).font = sum_font
        ws.cell(row=row, column=6).number_format = "#,##0.00 €"

        row += 1
        ws.cell(row=row, column=2, value="MwSt 19%:")
        ws.cell(row=row, column=6, value=float(lv.mwst))
        ws.cell(row=row, column=6).number_format = "#,##0.00 €"

        row += 1
        ws.cell(row=row, column=2, value="BRUTTO SUMME:").font = sum_font
        cell = ws.cell(row=row, column=6, value=float(lv.brutto_summe))
        cell.font = sum_font
        cell.fill = sum_fill
        cell.number_format = "#,##0.00 €"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _write_gruppe_excel(self, ws, gruppe: LosGruppe, row: int, sum_font, sum_fill) -> int:
        """Schreibt Gruppe in Excel"""
        from openpyxl.styles import Font, PatternFill

        # Gruppenzeile
        ws.cell(row=row, column=1, value=gruppe.oz).font = Font(bold=True)
        ws.cell(row=row, column=2, value=gruppe.bezeichnung).font = Font(bold=True)
        ws.cell(row=row, column=6, value=float(gruppe.summe)).font = Font(bold=True)
        ws.cell(row=row, column=6).number_format = "#,##0.00 €"
        row += 1

        # Positionen
        for pos in gruppe.positionen:
            ws.cell(row=row, column=1, value=pos.oz)
            ws.cell(row=row, column=2, value=pos.kurztext)
            ws.cell(row=row, column=3, value=float(pos.menge))
            ws.cell(row=row, column=4, value=pos.einheit.value)
            ws.cell(row=row, column=5, value=float(pos.einheitspreis))
            ws.cell(row=row, column=5).number_format = "#,##0.00"
            ws.cell(row=row, column=6, value=float(pos.gesamtpreis))
            ws.cell(row=row, column=6).number_format = "#,##0.00"
            row += 1

        # Untergruppen
        for ug in gruppe.untergruppen:
            row = self._write_gruppe_excel(ws, ug, row, sum_font, sum_fill)

        return row

    def _create_root(self, lv: Leistungsverzeichnis) -> ET.Element:
        """Erstellt GAEB XML Root"""
        root = ET.Element("GAEB")
        root.set("xmlns", self.GAEB_NAMESPACE)

        # GAEBInfo
        gaeb_info = ET.SubElement(root, "GAEBInfo")
        ET.SubElement(gaeb_info, "Version").text = "GAEB XML 3.2"
        ET.SubElement(gaeb_info, "VersNo").text = "32"
        ET.SubElement(gaeb_info, "Date").text = datetime.now().isoformat()
        ET.SubElement(gaeb_info, "ProgSystem").text = "IFC Dashboard"

        # PrjInfo
        prj_info = ET.SubElement(root, "PrjInfo")
        ET.SubElement(prj_info, "NamePrj").text = lv.projekt_name
        if lv.projekt_nummer:
            ET.SubElement(prj_info, "LblPrj").text = lv.projekt_nummer
        ET.SubElement(prj_info, "Cur").text = lv.waehrung

        # Award
        award = ET.SubElement(root, "Award")
        boq = ET.SubElement(award, "BoQ")
        ET.SubElement(boq, "BoQInfo")
        boq_body = ET.SubElement(boq, "BoQBody")

        for los in lv.lose:
            self._add_gruppe(boq_body, los)

        return root

    def _add_gruppe(self, parent: ET.Element, gruppe: LosGruppe):
        """Fügt Gruppe zum XML hinzu"""
        boq_ctgy = ET.SubElement(parent, "BoQCtgy")
        ET.SubElement(boq_ctgy, "LblTx").text = gruppe.oz
        ET.SubElement(boq_ctgy, "Headline").text = gruppe.bezeichnung

        boq_body = ET.SubElement(boq_ctgy, "BoQBody")

        for pos in gruppe.positionen:
            self._add_position(boq_body, pos)

        for ug in gruppe.untergruppen:
            self._add_gruppe(boq_body, ug)

    def _add_position(self, parent: ET.Element, pos: Position):
        """Fügt Position zum XML hinzu"""
        item = ET.SubElement(parent, "Itemlist")
        item_elem = ET.SubElement(item, "Item")

        ET.SubElement(item_elem, "Qty").text = str(pos.menge)
        ET.SubElement(item_elem, "QU").text = pos.einheit.value

        description = ET.SubElement(item_elem, "Description")
        ET.SubElement(description, "OutlineText").text = pos.kurztext
        if pos.langtext:
            ET.SubElement(description, "DetailTxt").text = pos.langtext

    def _indent(self, elem, level=0):
        """XML Einrückung"""
        i = "\n" + level * "  "
        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = i + "  "
            if not elem.tail or not elem.tail.strip():
                elem.tail = i
            for child in elem:
                self._indent(child, level + 1)
            if not child.tail or not child.tail.strip():
                child.tail = i
        else:
            if level and (not elem.tail or not elem.tail.strip()):
                elem.tail = i


class MassenermittlungHelper:
    """
    Hilfsklasse für Massenermittlung aus Räumen

    Erstellt LV-Positionen aus Raum-Daten
    """

    @staticmethod
    def from_rooms(
        rooms: List[Dict],
        gewerk: str,
        oz_prefix: str = "01",
        einheit: MengenEinheit = MengenEinheit.M2,
    ) -> List[Position]:
        """Erstellt Positionen aus Raumlisten (Bodenbeläge etc.)"""
        positionen = []

        for idx, room in enumerate(rooms, 1):
            pos = Position(
                oz=f"{oz_prefix}.01.{idx:04d}",
                kurztext=f"{gewerk} {room.get('number', '')} {room.get('name', '')}".strip(),
                menge=Decimal(str(room.get("area", 0))),
                einheit=einheit,
            )
            positionen.append(pos)

        return positionen

    @staticmethod
    def from_room_perimeters(
        rooms: List[Dict],
        gewerk: str = "Sockelleisten",
        oz_prefix: str = "02",
    ) -> List[Position]:
        """Erstellt Positionen aus Raumumfängen"""
        positionen = []

        for idx, room in enumerate(rooms, 1):
            perimeter = room.get("perimeter", 0)
            if perimeter > 0:
                pos = Position(
                    oz=f"{oz_prefix}.01.{idx:04d}",
                    kurztext=f"{gewerk} {room.get('number', '')} {room.get('name', '')}".strip(),
                    menge=Decimal(str(perimeter)),
                    einheit=MengenEinheit.M,
                )
                positionen.append(pos)

        return positionen

    @staticmethod
    def from_wall_areas(
        total_area: float,
        gewerk: str = "Wandanstrich",
        oz: str = "03.01.0010",
    ) -> Position:
        """Erstellt Position aus Wandflächen-Summe"""
        return Position(
            oz=oz,
            kurztext=gewerk,
            menge=Decimal(str(total_area)),
            einheit=MengenEinheit.M2,
        )
