# apps/cad_hub/views.py
"""
Views für IFC Dashboard
"""
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)

from .models import Door, Floor, IFCModel, IFCProject, Room, Slab, Wall, Window


class HtmxMixin:
    """Mixin für HTMX-Support: Liefert Partial bei HTMX-Request"""

    partial_template_name = None

    def get_template_names(self):
        if self.request.headers.get("HX-Request") and self.partial_template_name:
            return [self.partial_template_name]
        return super().get_template_names()


# =============================================================================
# Dashboard
# =============================================================================


class DashboardView(TemplateView):
    """Haupt-Dashboard mit Übersicht"""

    template_name = "cad_hub/dashboard.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["recent_projects"] = IFCProject.objects.all()[:5]
        ctx["stats"] = {
            "projects": IFCProject.objects.count(),
            "models": IFCModel.objects.filter(status="ready").count(),
            "rooms": Room.objects.count(),
        }

        return ctx


# =============================================================================
# Projekte
# =============================================================================


class ProjectListView(HtmxMixin, ListView):
    """Liste aller Projekte"""

    model = IFCProject
    template_name = "cad_hub/project_list.html"
    partial_template_name = "cad_hub/partials/_project_list.html"
    context_object_name = "projects"
    paginate_by = 10


class ProjectDetailView(DetailView):
    """Projekt-Detail mit Modellversionen"""

    model = IFCProject
    template_name = "cad_hub/project_detail.html"
    context_object_name = "project"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["models"] = self.object.models.all()
        return ctx


class ProjectCreateView(LoginRequiredMixin, CreateView):
    """Neues Projekt erstellen"""

    model = IFCProject
    template_name = "cad_hub/project_form.html"
    fields = ["name"]
    success_url = reverse_lazy("cad_hub:project_list")

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class ProjectUpdateView(LoginRequiredMixin, UpdateView):
    """Projekt-Name bearbeiten"""

    model = IFCProject
    template_name = "cad_hub/project_form.html"
    fields = ["name"]

    def get_success_url(self):
        return reverse_lazy("cad_hub:project_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, f'Projekt "{form.instance.name}" erfolgreich aktualisiert.')
        return super().form_valid(form)


class ProjectDeleteView(LoginRequiredMixin, DeleteView):
    """Projekt löschen (inkl. aller Modelle)"""

    model = IFCProject
    template_name = "cad_hub/project_confirm_delete.html"
    success_url = reverse_lazy("cad_hub:project_list")

    def delete(self, request, *args, **kwargs):
        project = self.get_object()
        messages.success(
            request, f'Projekt "{project.name}" und alle zugehörigen IFC-Versionen wurden gelöscht.'
        )
        return super().delete(request, *args, **kwargs)


# =============================================================================
# Modelle
# =============================================================================


class ModelDetailView(DetailView):
    """IFC-Modell Detail"""

    model = IFCModel
    template_name = "cad_hub/model_detail.html"
    context_object_name = "model"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        model = self.get_object()

        # Geschosse mit Raumanzahl
        floors = model.floors.annotate(room_count=Count("rooms")).order_by("sort_order")

        ctx["floors"] = floors
        ctx["room_count"] = model.rooms.count()
        return ctx


class ModelViewerView(DetailView):
    """3D Viewer für IFC-Modell"""

    model = IFCModel
    template_name = "cad_hub/model_viewer.html"
    context_object_name = "model"


class IFCContentOverviewView(DetailView):
    """IFC Inhalts-Übersicht: Alle extrahierten Elemente in Tabellen"""

    model = IFCModel
    template_name = "cad_hub/ifc_content_overview.html"
    context_object_name = "model"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        model = self.object

        # Aggregierte Statistiken
        ctx["stats"] = {
            "floors": model.floors.count(),
            "rooms": model.rooms.count(),
            "windows": model.windows.count(),
            "doors": model.doors.count(),
            "walls": model.walls.count(),
            "slabs": model.slabs.count(),
            # Flächen
            "total_room_area": model.rooms.aggregate(Sum("area"))["area__sum"] or 0,
            "total_wall_gross_area": model.walls.aggregate(Sum("gross_area"))["gross_area__sum"]
            or 0,
            "total_wall_net_area": model.walls.aggregate(Sum("net_area"))["net_area__sum"] or 0,
            "total_slab_area": model.slabs.aggregate(Sum("area"))["area__sum"] or 0,
            # Wände
            "external_walls": model.walls.filter(is_external=True).count(),
            "internal_walls": model.walls.filter(is_external=False).count(),
        }

        # Pro Geschoss
        ctx["floors_with_stats"] = []
        for floor in model.floors.all():
            ctx["floors_with_stats"].append(
                {
                    "floor": floor,
                    "rooms": floor.rooms.count(),
                    "windows": floor.windows.count(),
                    "doors": floor.doors.count(),
                    "walls": floor.walls.count(),
                    "slabs": floor.slabs.count(),
                    "room_area": floor.rooms.aggregate(Sum("area"))["area__sum"] or 0,
                }
            )

        # Beispieldaten (erste 5 pro Typ)
        ctx["sample_rooms"] = model.rooms.all()[:5]
        ctx["sample_windows"] = model.windows.all()[:5]
        ctx["sample_doors"] = model.doors.all()[:5]
        ctx["sample_walls"] = model.walls.all()[:5]
        ctx["sample_slabs"] = model.slabs.all()[:5]

        return ctx


class ModelUploadView(CreateView):
    """IFC-Datei hochladen"""

    model = IFCModel
    template_name = "cad_hub/model_upload.html"
    fields = ["ifc_file"]

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["project"] = get_object_or_404(IFCProject, pk=self.kwargs["project_id"])
        return ctx

    def form_valid(self, form):
        project = get_object_or_404(IFCProject, pk=self.kwargs["project_id"])

        # Version ermitteln
        last = IFCModel.objects.filter(project=project).order_by("-version").first()

        form.instance.project = project
        form.instance.version = (last.version + 1) if last else 1
        form.instance.status = IFCModel.Status.UPLOADING

        response = super().form_valid(form)

        # Processing starten (später async)
        from .tasks import process_ifc_upload

        process_ifc_upload(str(self.object.pk))

        return response

    def get_success_url(self):
        return reverse_lazy("cad_hub:model_detail", kwargs={"pk": self.object.pk})


class CADUploadView(LoginRequiredMixin, TemplateView):
    """DWG/DXF/PDF/GAEB Upload mit Multi-Format Support"""

    template_name = "cad_hub/cad_upload.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["project"] = get_object_or_404(IFCProject, pk=self.kwargs["project_id"])
        ctx["supported_formats"] = {
            "dwg": {"name": "AutoCAD DWG", "icon": "📐", "description": "AutoCAD Zeichnung"},
            "dxf": {"name": "AutoCAD DXF", "icon": "📐", "description": "Drawing Exchange Format"},
            "pdf": {"name": "PDF Plan", "icon": "📄", "description": "Gescannte Baupläne"},
            "x83": {
                "name": "GAEB X83",
                "icon": "📋",
                "description": "Leistungsverzeichnis (Angebot)",
            },
        }
        return ctx

    def post(self, request, *args, **kwargs):
        project = get_object_or_404(IFCProject, pk=self.kwargs["project_id"])
        uploaded_file = request.FILES.get("file")

        if not uploaded_file:
            messages.error(request, "Keine Datei ausgewählt.")
            return redirect("cad_hub:cad_upload", project_id=project.pk)

        # Dateiendung prüfen
        file_ext = uploaded_file.name.split(".")[-1].lower()

        if file_ext not in ["dwg", "dxf", "pdf", "x83", "x84"]:
            messages.error(request, f"Format .{file_ext} wird nicht unterstützt.")
            return redirect("cad_hub:cad_upload", project_id=project.pk)

        # Placeholder: In Zukunft MCP Backend Integration
        messages.info(
            request,
            f"Upload erfolgreich! Format: {file_ext.upper()} - "
            f"Konvertierung zu IFC folgt in zukünftiger Version.",
        )

        return redirect("cad_hub:project_detail", pk=project.pk)


class ModelDeleteView(LoginRequiredMixin, DeleteView):
    """IFC-Version löschen"""

    model = IFCModel
    template_name = "cad_hub/model_confirm_delete.html"

    def get_success_url(self):
        return reverse_lazy("cad_hub:project_detail", kwargs={"pk": self.object.project.pk})

    def delete(self, request, *args, **kwargs):
        model = self.get_object()
        project_id = model.project.pk
        messages.success(request, f"IFC-Version {model.version} wurde gelöscht.")
        return super().delete(request, *args, **kwargs)


# =============================================================================
# Räume
# =============================================================================


class RoomListView(HtmxMixin, ListView):
    """Raumliste mit Filterung"""

    model = Room
    template_name = "cad_hub/room_list.html"
    partial_template_name = "cad_hub/partials/_room_table.html"
    context_object_name = "rooms"
    paginate_by = 20

    def get_queryset(self):
        model_id = self.kwargs["model_id"]
        qs = Room.objects.filter(ifc_model_id=model_id)

        # Filter: Geschoss
        if floor := self.request.GET.get("floor"):
            qs = qs.filter(floor_id=floor)

        # Filter: Nutzung
        if usage := self.request.GET.get("usage"):
            qs = qs.filter(usage_category=usage)

        # Suche
        if search := self.request.GET.get("q"):
            qs = qs.filter(name__icontains=search)

        return qs.select_related("floor").order_by("number")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        model_id = self.kwargs["model_id"]

        ctx["ifc_model"] = get_object_or_404(IFCModel, pk=model_id)
        ctx["floors"] = Floor.objects.filter(ifc_model_id=model_id)
        ctx["usage_choices"] = Room.UsageCategory.choices

        return ctx


class RoomDetailView(HtmxMixin, DetailView):
    """Raum-Detail (für Seitenpanel)"""

    model = Room
    template_name = "cad_hub/room_detail.html"
    partial_template_name = "cad_hub/partials/_room_detail.html"
    context_object_name = "room"


# =============================================================================
# Flächen
# =============================================================================


class AreaSummaryView(HtmxMixin, TemplateView):
    """DIN 277 Flächenübersicht"""

    template_name = "cad_hub/area_summary.html"
    partial_template_name = "cad_hub/partials/_area_summary.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        model_id = self.kwargs["model_id"]

        # Get model and project reference
        ifc_model = get_object_or_404(IFCModel, pk=model_id)

        # Einfache Flächenberechnung aus Räumen
        rooms = Room.objects.filter(ifc_model=ifc_model)
        total_area = rooms.aggregate(Sum("area"))["area__sum"] or 0

        ctx["areas"] = {
            "bgf": total_area,
            "ngf": total_area * 0.85,  # Beispiel: 85% als Nutzfläche
            "nf": total_area * 0.75,
            "tf": total_area * 0.10,
            "vf": total_area * 0.05,
        }
        ctx["din277"] = {
            "total_area": total_area,
            "rooms_count": rooms.count(),
        }
        ctx["ifc_model"] = ifc_model

        return ctx


# =============================================================================
# Export
# =============================================================================


class ExportRaumbuchView(View):
    """Raumbuch als Excel exportieren"""

    def get(self, request, model_id):
        ifc_model = get_object_or_404(IFCModel, pk=model_id)
        export_type = request.GET.get("type", "raumbuch")

        from .services.export_service import RaumbuchExportService

        service = RaumbuchExportService()

        if export_type == "din277":
            output = service.export_din277_summary(ifc_model)
            filename = f"DIN277_{ifc_model.project.name}_v{ifc_model.version}.xlsx"
        else:
            output = service.export_to_excel(ifc_model)
            filename = f"Raumbuch_{ifc_model.project.name}_v{ifc_model.version}.xlsx"

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response


class ExportWoFlVView(View):
    """WoFlV Wohnflächenberechnung als Excel exportieren"""

    def get(self, request, model_id):
        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        ifc_model = get_object_or_404(IFCModel, pk=model_id)

        # WoFlV berechnen
        from .services import WoFlVCalculator

        rooms = list(
            Room.objects.filter(ifc_model=ifc_model).values("name", "number", "area", "height")
        )

        calculator = WoFlVCalculator()
        result = calculator.calculate_from_rooms(rooms)

        # Excel erstellen
        wb = Workbook()
        ws = wb.active
        ws.title = "WoFlV Berechnung"

        # Header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")

        ws["A1"] = f"WoFlV Wohnflächenberechnung"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = ifc_model.project.name

        # Zusammenfassung
        ws["A4"] = "Zusammenfassung"
        ws["A4"].font = Font(bold=True, size=12)

        summary_data = [
            ("Grundfläche gesamt:", float(result.grundflaeche_gesamt), "m²"),
            ("Wohnfläche 100%:", float(result.wohnflaeche_100), "m²"),
            ("Wohnfläche 50%:", float(result.wohnflaeche_50), "m²"),
            ("Wohnfläche 25%:", float(result.wohnflaeche_25), "m²"),
            ("Nicht angerechnet:", float(result.nicht_angerechnet), "m²"),
            ("", "", ""),
            ("WOHNFLÄCHE GESAMT:", float(result.wohnflaeche_gesamt), "m²"),
            ("Anrechnungsquote:", result.anrechnungsquote * 100, "%"),
        ]

        for idx, (label, value, unit) in enumerate(summary_data, 5):
            ws.cell(row=idx, column=1, value=label)
            if value:
                ws.cell(row=idx, column=2, value=value).number_format = "#,##0.00"
            ws.cell(row=idx, column=3, value=unit)

        # Raumdetails
        ws["A15"] = "Raumdetails"
        ws["A15"].font = Font(bold=True, size=12)

        headers = ["Nr.", "Raumname", "Grundfläche", "Höhe", "Typ", "Faktor", "Wohnfläche"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=16, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, room in enumerate(result.rooms, 17):
            ws.cell(row=row_idx, column=1, value=room.number)
            ws.cell(row=row_idx, column=2, value=room.name)
            ws.cell(row=row_idx, column=3, value=float(room.grundflaeche)).number_format = (
                "#,##0.00"
            )
            ws.cell(row=row_idx, column=4, value=float(room.hoehe)).number_format = "0.00"
            ws.cell(row=row_idx, column=5, value=room.raumtyp)
            ws.cell(row=row_idx, column=6, value=float(room.gesamt_faktor)).number_format = "0%"
            ws.cell(row=row_idx, column=7, value=float(room.wohnflaeche)).number_format = "#,##0.00"

        # Spaltenbreiten
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 25

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"WoFlV_{ifc_model.project.name}_v{ifc_model.version}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class ExportGAEBView(View):
    """GAEB Leistungsverzeichnis exportieren"""

    def get(self, request, model_id):
        from decimal import Decimal

        ifc_model = get_object_or_404(IFCModel, pk=model_id)
        format_type = request.GET.get("format", "excel")  # excel oder xml

        from .services import (
            GAEBGenerator,
            Leistungsverzeichnis,
            LosGruppe,
            MassenermittlungHelper,
            MengenEinheit,
            Position,
        )

        # Räume laden
        rooms = list(
            Room.objects.filter(ifc_model=ifc_model).values("name", "number", "area", "perimeter")
        )

        # LV erstellen mit Massenermittlung
        lv = Leistungsverzeichnis(
            projekt_name=ifc_model.project.name,
            projekt_nummer=str(ifc_model.project.pk)[:8],
        )

        # Los 1: Bodenbeläge
        boden_positionen = MassenermittlungHelper.from_rooms(
            rooms, gewerk="Bodenbelag", oz_prefix="01"
        )
        lv.lose.append(LosGruppe(oz="01", bezeichnung="Bodenbeläge", positionen=boden_positionen))

        # Los 2: Sockelleisten
        sockel_positionen = MassenermittlungHelper.from_room_perimeters(
            rooms, gewerk="Sockelleisten", oz_prefix="02"
        )
        if sockel_positionen:
            lv.lose.append(
                LosGruppe(oz="02", bezeichnung="Sockelleisten", positionen=sockel_positionen)
            )

        # Export
        generator = GAEBGenerator()

        if format_type == "xml":
            output = generator.generate_xml(lv)
            content_type = "application/xml"
            filename = f"LV_{ifc_model.project.name}_v{ifc_model.version}.x84"
        else:
            output = generator.generate_excel(lv)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"LV_{ifc_model.project.name}_v{ifc_model.version}.xlsx"

        response = HttpResponse(output.read(), content_type=content_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class ExportX83View(View):
    """
    IFC → GAEB X83 Export (Angebot mit Mengen und Preisen)
    
    Extrahiert alle Mengen aus dem IFC-Modell und erstellt
    ein vollständiges Leistungsverzeichnis nach GAEB X83.
    
    Query-Parameter:
        format: xml (default) oder excel
        gewerke: kommaseparierte Liste (z.B. bodenbelag,tueren,fenster)
        prices: 1/0 - Einheitspreise inkludieren
    """

    def get(self, request, model_id):
        ifc_model = get_object_or_404(IFCModel, pk=model_id)
        
        format_type = request.GET.get("format", "xml")
        include_prices = request.GET.get("prices", "1") == "1"
        gewerke_param = request.GET.get("gewerke", "")
        
        selected_gewerke = None
        if gewerke_param:
            selected_gewerke = [g.strip() for g in gewerke_param.split(",")]
        
        from .services import get_ifc_x83_converter
        
        # IFC-Daten aus Datenbank laden
        ifc_data = self._extract_ifc_data(ifc_model)
        
        # Konvertieren
        converter = get_ifc_x83_converter()
        
        if format_type == "excel":
            output = converter.convert_to_excel(
                ifc_data=ifc_data,
                projekt_name=ifc_model.project.name,
                projekt_nummer=str(ifc_model.project.pk)[:8],
                include_prices=include_prices,
                selected_gewerke=selected_gewerke,
            )
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"LV_X83_{ifc_model.project.name}_v{ifc_model.version}.xlsx"
        else:
            output = converter.convert_to_x83(
                ifc_data=ifc_data,
                projekt_name=ifc_model.project.name,
                projekt_nummer=str(ifc_model.project.pk)[:8],
                include_prices=include_prices,
                selected_gewerke=selected_gewerke,
            )
            content_type = "application/xml"
            filename = f"LV_{ifc_model.project.name}_v{ifc_model.version}.x83"
        
        response = HttpResponse(output.read(), content_type=content_type)
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
    
    def _extract_ifc_data(self, ifc_model) -> dict:
        """Extrahiert IFC-Daten aus der Datenbank."""
        rooms = list(
            Room.objects.filter(ifc_model=ifc_model).values(
                "name", "number", "area", "perimeter", "height", "volume"
            )
        )
        
        walls = list(
            Wall.objects.filter(ifc_model=ifc_model).values(
                "name", "ifc_guid", "length", "height", "thickness"
            )
        )
        # Wandfläche berechnen
        for wall in walls:
            wall["area"] = (wall.get("length", 0) or 0) * (wall.get("height", 0) or 0)
        
        doors = list(
            Door.objects.filter(ifc_model=ifc_model).values(
                "name", "ifc_guid", "width", "height"
            )
        )
        # Türtyp aus Name extrahieren
        for door in doors:
            door["type"] = "Standard"
            if "brand" in (door.get("name", "") or "").lower():
                door["type"] = "Brandschutz"
        
        windows = list(
            Window.objects.filter(ifc_model=ifc_model).values(
                "name", "ifc_guid", "width", "height"
            )
        )
        
        slabs = list(
            Slab.objects.filter(ifc_model=ifc_model).values(
                "name", "ifc_guid", "area", "thickness"
            )
        )
        
        return {
            "rooms": rooms,
            "walls": walls,
            "doors": doors,
            "windows": windows,
            "slabs": slabs,
        }


class WoFlVSummaryView(HtmxMixin, TemplateView):
    """WoFlV Wohnflächenübersicht"""

    template_name = "cad_hub/woflv_summary.html"
    partial_template_name = "cad_hub/partials/_woflv_summary.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        model_id = self.kwargs["model_id"]

        # Get model and project reference
        ifc_model = get_object_or_404(IFCModel, pk=model_id)

        # Einfache WoFlV-Berechnung aus Räumen
        rooms = Room.objects.filter(ifc_model=ifc_model)

        ctx["woflv"] = {
            "wohnflaeche_gesamt": 0,
            "grundflaeche_gesamt": 0,
        }
        ctx["woflv_rooms"] = []
        ctx["ifc_model"] = ifc_model

        return ctx


# =============================================================================
# Fenster, Türen, Wände, Decken
# =============================================================================


class WindowListView(HtmxMixin, ListView):
    """Fensterliste mit Filterung"""

    model = Window
    template_name = "cad_hub/window_list.html"
    partial_template_name = "cad_hub/partials/_window_table.html"
    context_object_name = "windows"
    paginate_by = 50

    def get_queryset(self):
        model_id = self.kwargs["model_id"]
        qs = Window.objects.filter(ifc_model_id=model_id).select_related("floor", "room")

        if floor_id := self.request.GET.get("floor"):
            qs = qs.filter(floor_id=floor_id)
        if room_id := self.request.GET.get("room"):
            qs = qs.filter(room_id=room_id)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        model = get_object_or_404(IFCModel, pk=self.kwargs["model_id"])
        ctx["model"] = model
        ctx["floors"] = model.floors.all()
        ctx["total_count"] = self.get_queryset().count()
        ctx["total_area"] = self.get_queryset().aggregate(Sum("area"))["area__sum"] or 0
        return ctx


class DoorListView(HtmxMixin, ListView):
    """Türliste mit Filterung"""

    model = Door
    template_name = "cad_hub/door_list.html"
    partial_template_name = "cad_hub/partials/_door_table.html"
    context_object_name = "doors"
    paginate_by = 50

    def get_queryset(self):
        model_id = self.kwargs["model_id"]
        qs = Door.objects.filter(ifc_model_id=model_id).select_related(
            "floor", "from_room", "to_room"
        )

        if floor_id := self.request.GET.get("floor"):
            qs = qs.filter(floor_id=floor_id)
        if fire_rating := self.request.GET.get("fire_rating"):
            qs = qs.filter(fire_rating=fire_rating)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        model = get_object_or_404(IFCModel, pk=self.kwargs["model_id"])
        ctx["model"] = model
        ctx["floors"] = model.floors.all()
        ctx["total_count"] = self.get_queryset().count()
        return ctx


class WallListView(HtmxMixin, ListView):
    """Wandliste mit Filterung"""

    model = Wall
    template_name = "cad_hub/wall_list.html"
    partial_template_name = "cad_hub/partials/_wall_table.html"
    context_object_name = "walls"
    paginate_by = 50

    def get_queryset(self):
        model_id = self.kwargs["model_id"]
        qs = Wall.objects.filter(ifc_model_id=model_id).select_related("floor")

        if floor_id := self.request.GET.get("floor"):
            qs = qs.filter(floor_id=floor_id)
        if wall_type := self.request.GET.get("type"):
            if wall_type == "external":
                qs = qs.filter(is_external=True)
            elif wall_type == "internal":
                qs = qs.filter(is_external=False)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        model = get_object_or_404(IFCModel, pk=self.kwargs["model_id"])
        ctx["model"] = model
        ctx["floors"] = model.floors.all()
        ctx["total_count"] = self.get_queryset().count()
        ctx["total_gross_area"] = (
            self.get_queryset().aggregate(Sum("gross_area"))["gross_area__sum"] or 0
        )
        ctx["total_net_area"] = self.get_queryset().aggregate(Sum("net_area"))["net_area__sum"] or 0
        ctx["external_count"] = self.get_queryset().filter(is_external=True).count()
        ctx["internal_count"] = self.get_queryset().filter(is_external=False).count()
        return ctx


class SlabListView(HtmxMixin, ListView):
    """Deckenliste mit Filterung"""

    model = Slab
    template_name = "cad_hub/slab_list.html"
    partial_template_name = "cad_hub/partials/_slab_table.html"
    context_object_name = "slabs"
    paginate_by = 50

    def get_queryset(self):
        model_id = self.kwargs["model_id"]
        qs = Slab.objects.filter(ifc_model_id=model_id).select_related("floor")

        if floor_id := self.request.GET.get("floor"):
            qs = qs.filter(floor_id=floor_id)
        if slab_type := self.request.GET.get("slab_type"):
            qs = qs.filter(slab_type=slab_type)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        model = get_object_or_404(IFCModel, pk=self.kwargs["model_id"])
        ctx["model"] = model
        ctx["floors"] = model.floors.all()
        ctx["total_count"] = self.get_queryset().count()
        ctx["total_area"] = self.get_queryset().aggregate(Sum("area"))["area__sum"] or 0
        ctx["total_volume"] = self.get_queryset().aggregate(Sum("volume"))["volume__sum"] or 0
        return ctx
