"""Excel Export Service.

Generates Excel reports for room books, element lists, and DIN 277 calculations.
"""

from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@dataclass
class ExportConfig:
    """Configuration for Excel export."""

    title: str = "CAD-Hub Export"
    author: str = "CAD-Hub"
    company: str = ""
    include_header: bool = True
    include_summary: bool = True
    date_format: str = "%d.%m.%Y"


@dataclass
class RoomData:
    """Room data for export."""

    room_id: int
    name: str
    number: str
    floor_name: str
    area_m2: Decimal
    usage_type: str = ""
    din277_category: str = ""
    height_m: Decimal | None = None
    volume_m3: Decimal | None = None
    properties: dict[str, Any] = field(default_factory=dict)


@dataclass
class ElementData:
    """Element data for export."""

    element_id: int
    ifc_guid: str
    element_type: str
    name: str
    floor_name: str
    material: str = ""
    fire_rating: str = ""
    properties: dict[str, Any] = field(default_factory=dict)


class ExcelExportService:
    """Service for generating Excel exports."""

    # Style definitions
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79") if OPENPYXL_AVAILABLE else None
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11) if OPENPYXL_AVAILABLE else None
    TITLE_FONT = Font(bold=True, size=14) if OPENPYXL_AVAILABLE else None
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    ) if OPENPYXL_AVAILABLE else None

    def __init__(self, config: ExportConfig | None = None):
        """Initialize service.

        Args:
            config: Export configuration
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )

        self.config = config or ExportConfig()

    def export_room_book(
        self,
        rooms: list[RoomData],
        project_name: str,
        output_path: Path | None = None,
    ) -> bytes | None:
        """Export room book (Raumbuch) to Excel.

        Args:
            rooms: List of room data
            project_name: Project name for header
            output_path: Optional file path

        Returns:
            Excel bytes if no output_path, else None
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Raumbuch"

        row = 1

        # Title
        if self.config.include_header:
            ws.cell(row, 1, f"Raumbuch - {project_name}").font = self.TITLE_FONT
            row += 1
            ws.cell(row, 1, f"Erstellt: {datetime.now().strftime(self.config.date_format)}")
            row += 2

        # Headers
        headers = [
            "Nr.", "Raum-Nr.", "Raumname", "Etage", "Fläche (m²)",
            "Nutzungsart", "DIN 277", "Höhe (m)", "Volumen (m³)",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center")

        row += 1

        # Data rows
        total_area = Decimal("0")
        for i, room in enumerate(rooms, 1):
            ws.cell(row, 1, i).border = self.BORDER
            ws.cell(row, 2, room.number).border = self.BORDER
            ws.cell(row, 3, room.name).border = self.BORDER
            ws.cell(row, 4, room.floor_name).border = self.BORDER

            area_cell = ws.cell(row, 5, float(room.area_m2))
            area_cell.number_format = "#,##0.00"
            area_cell.border = self.BORDER

            ws.cell(row, 6, room.usage_type).border = self.BORDER
            ws.cell(row, 7, room.din277_category).border = self.BORDER

            if room.height_m:
                height_cell = ws.cell(row, 8, float(room.height_m))
                height_cell.number_format = "0.00"
                height_cell.border = self.BORDER

            if room.volume_m3:
                vol_cell = ws.cell(row, 9, float(room.volume_m3))
                vol_cell.number_format = "#,##0.00"
                vol_cell.border = self.BORDER

            total_area += room.area_m2
            row += 1

        # Summary
        if self.config.include_summary:
            row += 1
            ws.cell(row, 4, "Gesamt:").font = Font(bold=True)
            total_cell = ws.cell(row, 5, float(total_area))
            total_cell.number_format = "#,##0.00"
            total_cell.font = Font(bold=True)

        # Column widths
        widths = [6, 12, 25, 15, 12, 20, 10, 10, 12]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        return self._save_workbook(wb, output_path)

    def export_din277(
        self,
        floors: list[dict[str, Any]],
        project_name: str,
        output_path: Path | None = None,
    ) -> bytes | None:
        """Export DIN 277 calculation to Excel.

        Args:
            floors: Floor data with area calculations
            project_name: Project name
            output_path: Optional file path

        Returns:
            Excel bytes if no output_path
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "DIN 277"

        row = 1

        # Title
        if self.config.include_header:
            ws.cell(row, 1, f"DIN 277 Flächenberechnung - {project_name}").font = self.TITLE_FONT
            row += 1
            ws.cell(row, 1, f"Erstellt: {datetime.now().strftime(self.config.date_format)}")
            row += 2

        # Headers
        headers = ["Etage", "BGF (m²)", "KGF (m²)", "NGF (m²)", "NUF (m²)", "TF (m²)", "VF (m²)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center")

        row += 1

        # Data
        totals = {"bgf": 0, "kgf": 0, "ngf": 0, "nuf": 0, "tf": 0, "vf": 0}
        for floor in floors:
            ws.cell(row, 1, floor.get("name", "")).border = self.BORDER

            for i, key in enumerate(["bgf", "kgf", "ngf", "nuf", "tf", "vf"], 2):
                value = floor.get(key, 0)
                cell = ws.cell(row, i, float(value))
                cell.number_format = "#,##0.00"
                cell.border = self.BORDER
                totals[key] += value

            row += 1

        # Totals
        row += 1
        ws.cell(row, 1, "GESAMT").font = Font(bold=True)
        for i, key in enumerate(["bgf", "kgf", "ngf", "nuf", "tf", "vf"], 2):
            cell = ws.cell(row, i, float(totals[key]))
            cell.number_format = "#,##0.00"
            cell.font = Font(bold=True)
            cell.border = self.BORDER

        # Legend
        row += 3
        ws.cell(row, 1, "Legende:").font = Font(bold=True)
        legends = [
            ("BGF", "Brutto-Grundfläche"),
            ("KGF", "Konstruktions-Grundfläche"),
            ("NGF", "Netto-Grundfläche"),
            ("NUF", "Nutzungsfläche"),
            ("TF", "Technikfläche"),
            ("VF", "Verkehrsfläche"),
        ]
        for abbr, desc in legends:
            row += 1
            ws.cell(row, 1, abbr).font = Font(bold=True)
            ws.cell(row, 2, desc)

        # Column widths
        widths = [15, 12, 12, 12, 12, 12, 12]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        return self._save_workbook(wb, output_path)

    def export_element_list(
        self,
        elements: list[ElementData],
        project_name: str,
        element_type: str = "all",
        output_path: Path | None = None,
    ) -> bytes | None:
        """Export element list to Excel.

        Args:
            elements: List of element data
            project_name: Project name
            element_type: Filter by type or 'all'
            output_path: Optional file path

        Returns:
            Excel bytes if no output_path
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Elementliste"

        row = 1

        # Title
        type_label = element_type if element_type != "all" else "Alle Elemente"
        if self.config.include_header:
            ws.cell(row, 1, f"Elementliste ({type_label}) - {project_name}").font = self.TITLE_FONT
            row += 1
            ws.cell(row, 1, f"Erstellt: {datetime.now().strftime(self.config.date_format)}")
            row += 2

        # Headers
        headers = ["Nr.", "IFC GUID", "Typ", "Name", "Etage", "Material", "Brandschutz"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center")

        row += 1

        # Filter and write data
        filtered = [e for e in elements if element_type == "all" or e.element_type == element_type]
        for i, elem in enumerate(filtered, 1):
            ws.cell(row, 1, i).border = self.BORDER
            ws.cell(row, 2, elem.ifc_guid).border = self.BORDER
            ws.cell(row, 3, elem.element_type).border = self.BORDER
            ws.cell(row, 4, elem.name).border = self.BORDER
            ws.cell(row, 5, elem.floor_name).border = self.BORDER
            ws.cell(row, 6, elem.material).border = self.BORDER
            ws.cell(row, 7, elem.fire_rating).border = self.BORDER
            row += 1

        # Summary
        if self.config.include_summary:
            row += 1
            ws.cell(row, 1, f"Anzahl Elemente: {len(filtered)}").font = Font(bold=True)

        # Column widths
        widths = [6, 25, 12, 30, 15, 20, 12]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        return self._save_workbook(wb, output_path)

    def export_fire_safety_summary(
        self,
        rated_elements: list[dict[str, Any]],
        project_name: str,
        output_path: Path | None = None,
    ) -> bytes | None:
        """Export fire safety summary to Excel.

        Args:
            rated_elements: Fire-rated elements data
            project_name: Project name
            output_path: Optional file path

        Returns:
            Excel bytes if no output_path
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Brandschutz"

        row = 1

        # Title
        if self.config.include_header:
            ws.cell(row, 1, f"Brandschutz-Übersicht - {project_name}").font = self.TITLE_FONT
            row += 1
            ws.cell(row, 1, f"Erstellt: {datetime.now().strftime(self.config.date_format)}")
            row += 2

        # Headers
        headers = ["Nr.", "Typ", "Name", "IFC GUID", "Soll", "Ist", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER

        row += 1

        # Data
        compliant = 0
        non_compliant = 0
        ok_fill = PatternFill("solid", fgColor="C6EFCE")
        fail_fill = PatternFill("solid", fgColor="FFC7CE")

        for i, elem in enumerate(rated_elements, 1):
            is_ok = elem.get("is_compliant", True)
            fill = ok_fill if is_ok else fail_fill

            ws.cell(row, 1, i).border = self.BORDER
            ws.cell(row, 2, elem.get("element_type", "")).border = self.BORDER
            ws.cell(row, 3, elem.get("name", "")).border = self.BORDER
            ws.cell(row, 4, elem.get("ifc_guid", "")).border = self.BORDER
            ws.cell(row, 5, elem.get("required_rating", "-")).border = self.BORDER
            ws.cell(row, 6, elem.get("actual_rating", "-")).border = self.BORDER

            status_cell = ws.cell(row, 7, "OK" if is_ok else "FEHLT")
            status_cell.border = self.BORDER
            status_cell.fill = fill

            if is_ok:
                compliant += 1
            else:
                non_compliant += 1
            row += 1

        # Summary
        row += 2
        ws.cell(row, 1, "Zusammenfassung:").font = Font(bold=True)
        row += 1
        ws.cell(row, 1, "Geprüfte Elemente:")
        ws.cell(row, 2, len(rated_elements))
        row += 1
        ws.cell(row, 1, "Konform:")
        ws.cell(row, 2, compliant)
        row += 1
        ws.cell(row, 1, "Nicht konform:")
        ws.cell(row, 2, non_compliant)

        # Column widths
        widths = [6, 12, 25, 25, 10, 10, 10]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        return self._save_workbook(wb, output_path)

    def _save_workbook(self, wb: "Workbook", output_path: Path | None) -> bytes | None:
        """Save workbook to file or return bytes."""
        if output_path:
            wb.save(output_path)
            return None

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
